"""
Script to create the initial Excel template.
"""

import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook import Workbook

def create_template():
    """Create the initial Excel template with required sheets."""
    # Get project root
    project_root = Path(__file__).parent.parent
    template_dir = project_root / "data" / "templates"
    template_dir.mkdir(parents=True, exist_ok=True)
    
    # Create workbook with specific settings
    wb = Workbook(write_only=False)
    
    # Create Scenario Inputs sheet
    ws_scenario = wb.active
    ws_scenario.title = "Scenario Inputs"
    
    # Add headers
    headers = [
        ["Scenario Information", ""],
        ["Name", ""],
        ["Description", ""],
        ["Is Base Case?", "No"],
        ["", ""],
        ["Financial Parameters", ""],
        ["Initial Revenue", "0"],
        ["Initial Costs", "0"],
        ["Annual Revenue Growth (%)", "0"],
        ["Annual Cost Growth (%)", "0"],
        ["Debt Ratio (%)", "0"],
        ["Interest Rate (%)", "0"],
        ["Tax Rate (%)", "0"]
    ]
    
    for row in headers:
        ws_scenario.append(row)
    
    # Create Equipment Inputs sheet
    ws_equipment = wb.create_sheet("Equipment Inputs")
    
    # Add headers
    equipment_headers = [
        ["Equipment Information", "", "", "", "", "", "", "", "", ""],
        ["Name", "Cost", "Useful Life (years)", "Max Capacity", "Maintenance Cost (%)", 
         "Availability (%)", "Purchase Year", "Is Leased?", "Lease Rate (%)", "Lease Type"]
    ]
    
    for row in equipment_headers:
        ws_equipment.append(row)
    
    # Create Product Inputs sheet
    ws_product = wb.create_sheet("Product Inputs")
    
    # Add headers
    product_headers = [
        ["Product Information", "", "", "", "", "", ""],
        ["Name", "Initial Units", "Unit Price", "Growth Rate (%)", 
         "Introduction Year", "Market Size", "Price Elasticity"]
    ]
    
    for row in product_headers:
        ws_product.append(row)
    
    # Apply formatting
    for ws in [ws_scenario, ws_equipment, ws_product]:
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Format data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
    
    # Save template with specific settings
    template_file = template_dir / "scenario_template.xlsx"
    wb.save(template_file)
    
    print(f"Template created successfully at: {template_file}")

if __name__ == "__main__":
    create_template() 