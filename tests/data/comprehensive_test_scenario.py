"""
Script to generate a comprehensive test workbook covering all functionality.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from decimal import Decimal
from datetime import datetime, date

def create_comprehensive_test_workbook():
    wb = openpyxl.Workbook()
    
    # Styling
    header_style = {
        'font': Font(bold=True, size=12),
        'fill': PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # 1. Overview Sheet
    ws_overview = wb.active
    ws_overview.title = "Overview"
    
    ws_overview.append(["Manufacturing Expansion Scenario Analysis"])
    ws_overview.append(["Base Case - High Tech Manufacturing"])
    ws_overview.append([])
    
    ws_overview.append(["Analysis Options"])
    ws_overview.append(["Include Optimization", "Yes"])
    ws_overview.append(["Run Monte Carlo", "Yes"])
    ws_overview.append(["Sensitivity Analysis", "Yes"])
    ws_overview.append(["Generate Charts", "Yes"])
    
    # 2. Scenario Details
    ws_scenario = wb.create_sheet("Scenario Details")
    
    ws_scenario.append(["Basic Information"])
    ws_scenario.append(["Name", "High Tech Manufacturing Expansion"])
    ws_scenario.append(["Description", "Expansion scenario for high-tech manufacturing facility"])
    ws_scenario.append(["Base Case", "Yes"])
    ws_scenario.append([])
    
    ws_scenario.append(["Financial Assumptions"])
    ws_scenario.append(["Initial Revenue", 5000000])
    ws_scenario.append(["Initial Costs", 3750000])
    ws_scenario.append(["Annual Revenue Growth (%)", 8])
    ws_scenario.append(["Annual Cost Growth (%)", 6])
    ws_scenario.append(["Debt Ratio (%)", 40])
    ws_scenario.append(["Interest Rate (%)", 5])
    ws_scenario.append(["Tax Rate (%)", 21])
    ws_scenario.append([])
    
    ws_scenario.append(["Time Horizon"])
    ws_scenario.append(["Start Year", 2024])
    ws_scenario.append(["Projection Years", 5])
    
    # 3. Equipment Sheet
    ws_equipment = wb.create_sheet("Equipment")
    
    headers = [
        "Name", "Type", "Cost", "Useful Life", "Max Capacity",
        "Maintenance Cost (%)", "Availability (%)", "Purchase Year",
        "Leased", "Lease Rate (%)", "Lease Type",
        "Optimization Priority", "Min Units", "Max Units",
        "Required Training (hours)", "Energy Usage (kWh)",
        "Space Required (sq ft)", "Installation Cost"
    ]
    ws_equipment.append(headers)
    
    equipment_data = [
        ["CNC Machine A", "Production", 750000, 10, 2000, 5, 95, 2024, "No", "", "", "High", 1, 3, 40, 25, 200, 50000],
        ["CNC Machine B", "Production", 1000000, 12, 3000, 6, 93, 2024, "Yes", 8, "Operating", "High", 1, 2, 60, 35, 300, 75000],
        ["Quality System", "QA", 250000, 8, 5000, 3, 98, 2024, "No", "", "", "Medium", 1, 1, 20, 5, 100, 25000],
        ["Assembly Line", "Assembly", 1500000, 15, 4000, 7, 92, 2025, "No", "", "", "High", 1, 2, 80, 40, 500, 100000],
        ["Testing Equipment", "QA", 500000, 7, 3000, 4, 96, 2025, "Yes", 7, "Capital", "Medium", 1, 2, 30, 15, 150, 35000]
    ]
    
    for row in equipment_data:
        ws_equipment.append(row)
    
    # 4. Products Sheet
    ws_products = wb.create_sheet("Products")
    
    headers = [
        "Name", "Category", "Initial Units", "Unit Price",
        "Growth Rate (%)", "Introduction Year", "Market Size",
        "Price Elasticity", "Variable Cost", "Marketing Cost",
        "Required Equipment", "Min Production", "Max Production",
        "Quality Requirements", "Lead Time (days)", "Shelf Life (days)"
    ]
    ws_products.append(headers)
    
    products_data = [
        ["Product X1", "Premium", 2000, 500, 10, 2024, 10000, -1.2, 300, 50000, "CNC Machine A, Assembly Line", 500, 5000, "High", 5, 365],
        ["Product X2", "Standard", 3000, 300, 8, 2024, 15000, -1.5, 180, 30000, "CNC Machine B", 1000, 8000, "Medium", 3, 365],
        ["Product Y1", "Premium", 1500, 800, 12, 2024, 8000, -1.1, 500, 80000, "CNC Machine A, CNC Machine B", 300, 4000, "High", 7, 180],
        ["Product Y2", "Standard", 2500, 400, 7, 2025, 12000, -1.4, 250, 40000, "Assembly Line", 800, 6000, "Medium", 4, 180],
        ["Product Z", "Custom", 1000, 1200, 15, 2025, 5000, -0.9, 800, 100000, "CNC Machine A, CNC Machine B, Assembly Line", 200, 3000, "High", 10, 90]
    ]
    
    for row in products_data:
        ws_products.append(row)
    
    # 5. Cost Drivers Sheet
    ws_costs = wb.create_sheet("Cost Drivers")
    
    headers = [
        "Category", "Type", "Base Cost", "Growth Rate (%)",
        "Start Year", "End Year", "Fixed Component",
        "Variable Component", "Step Costs", "Optimization Priority"
    ]
    ws_costs.append(headers)
    
    costs_data = [
        ["Labor", "Semi-Variable", 1500000, 4, 2024, 2028, 800000, 700000, "{'1000': 100000, '2000': 200000}", "High"],
        ["Materials", "Variable", 2000000, 3, 2024, 2028, 0, 2000000, "", "High"],
        ["Energy", "Semi-Variable", 300000, 5, 2024, 2028, 100000, 200000, "{'1500': 50000}", "Medium"],
        ["Maintenance", "Fixed", 250000, 3, 2024, 2028, 250000, 0, "", "Medium"],
        ["G&A", "Fixed", 500000, 2, 2024, 2028, 500000, 0, "", "Low"]
    ]
    
    for row in costs_data:
        ws_costs.append(row)
    
    # 6. Market Analysis Sheet
    ws_market = wb.create_sheet("Market Analysis")
    
    headers = [
        "Segment", "Growth Rate (%)", "Market Size",
        "Competition Level", "Entry Barriers", "Price Sensitivity",
        "Quality Sensitivity", "Lead Time Sensitivity"
    ]
    ws_market.append(headers)
    
    market_data = [
        ["Premium", 12, 25000, "High", "High", "Low", "High", "Medium"],
        ["Standard", 8, 50000, "Medium", "Medium", "High", "Medium", "High"],
        ["Custom", 15, 15000, "Low", "High", "Low", "High", "Low"]
    ]
    
    for row in market_data:
        ws_market.append(row)
    
    # 7. Optimization Parameters
    ws_optimization = wb.create_sheet("Optimization")
    
    ws_optimization.append(["Optimization Parameters"])
    ws_optimization.append(["Parameter", "Value", "Min", "Max"])
    ws_optimization.append(["Budget Constraint", 5000000, 0, 10000000])
    ws_optimization.append(["Minimum ROI (%)", 15, 10, 25])
    ws_optimization.append(["Maximum Debt Ratio (%)", 50, 30, 60])
    ws_optimization.append(["Target Utilization (%)", 85, 70, 95])
    ws_optimization.append([])
    
    ws_optimization.append(["Monte Carlo Parameters"])
    ws_optimization.append(["Parameter", "Value"])
    ws_optimization.append(["Number of Iterations", 1000])
    ws_optimization.append(["Confidence Level (%)", 95])
    ws_optimization.append([])
    
    ws_optimization.append(["Sensitivity Analysis"])
    ws_optimization.append(["Variable", "Min (%)", "Max (%)", "Step (%)"])
    ws_optimization.append(["Revenue Growth", -20, 20, 5])
    ws_optimization.append(["Cost Growth", -15, 15, 5])
    ws_optimization.append(["Market Size", -30, 30, 10])
    
    # Apply formatting
    for ws in wb.worksheets:
        # Format headers
        for cell in ws[1]:
            for key, value in header_style.items():
                setattr(cell, key, value)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Add data validation
    # Equipment sheet validations
    lease_types = ["Operating", "Capital"]
    priorities = ["High", "Medium", "Low"]
    
    dv_lease = DataValidation(type="list", formula1=f'"{",".join(lease_types)}"', allow_blank=True)
    dv_priority = DataValidation(type="list", formula1=f'"{",".join(priorities)}"', allow_blank=True)
    
    ws_equipment.add_data_validation(dv_lease)
    ws_equipment.add_data_validation(dv_priority)
    
    dv_lease.add(f"K2:K{ws_equipment.max_row}")
    dv_priority.add(f"L2:L{ws_equipment.max_row}")
    
    # Save workbook
    wb.save("tests/data/comprehensive_test_scenario.xlsx")

if __name__ == "__main__":
    create_comprehensive_test_workbook() 