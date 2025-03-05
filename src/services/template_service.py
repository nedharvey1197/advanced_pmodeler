"""
Template service for handling Excel template-based scenario creation and analysis.
"""

import pandas as pd
import openpyxl
from typing import Dict, Any, Optional
from datetime import datetime
import os
from pathlib import Path

from .spreadsheet_services import SpreadsheetService
from ..models import (
    Scenario, Equipment, Product, CostDriver, FinancialProjection,
    get_session
)

class TemplateService:
    """
    Service for handling Excel template-based workflow.
    
    This service provides methods for:
    - Creating scenario templates
    - Reading filled templates
    - Generating analysis from templates
    - Managing template versions
    """
    
    def __init__(self, session):
        """
        Initialize TemplateService with a database session.
        
        Args:
            session: SQLAlchemy database session
        """
        self.session = session
        self.spreadsheet_service = SpreadsheetService(session)
    
    def read_template(self, template_file: str) -> Dict[str, Any]:
        """
        Read a filled template and extract scenario data.
        
        Args:
            template_file: Path to the filled template
            
        Returns:
            Dictionary containing scenario data
        """
        try:
            wb = openpyxl.load_workbook(template_file, data_only=True)
            
            # Read scenario inputs
            ws_scenario = wb["Scenario Inputs"]
            scenario_data = {
                "name": ws_scenario["B2"].value,
                "description": ws_scenario["B3"].value,
                "is_base_case": ws_scenario["B4"].value.lower() == "yes",
                "initial_revenue": float(ws_scenario["B7"].value or 0),
                "initial_costs": float(ws_scenario["B8"].value or 0),
                "annual_revenue_growth": float(ws_scenario["B9"].value or 0) / 100,
                "annual_cost_growth": float(ws_scenario["B10"].value or 0) / 100,
                "debt_ratio": float(ws_scenario["B11"].value or 0) / 100,
                "interest_rate": float(ws_scenario["B12"].value or 0) / 100,
                "tax_rate": float(ws_scenario["B13"].value or 0) / 100
            }
            
            # Read equipment inputs
            ws_equipment = wb["Equipment Inputs"]
            equipment_data = []
            for row in ws_equipment.iter_rows(min_row=2):
                if row[0].value:  # If name is not empty
                    equipment_data.append({
                        "name": row[0].value,
                        "cost": float(row[1].value or 0),
                        "useful_life": int(row[2].value or 0),
                        "max_capacity": float(row[3].value or 0),
                        "maintenance_cost_pct": float(row[4].value or 0) / 100,
                        "availability_pct": float(row[5].value or 0) / 100,
                        "purchase_year": int(row[6].value or 0),
                        "is_leased": row[7].value.lower() == "yes",
                        "lease_rate": float(row[8].value or 0) / 100 if row[8].value else None,
                        "lease_type": row[9].value
                    })
            
            # Read product inputs
            ws_product = wb["Product Inputs"]
            product_data = []
            for row in ws_product.iter_rows(min_row=2):
                if row[0].value:  # If name is not empty
                    product_data.append({
                        "name": row[0].value,
                        "initial_units": float(row[1].value or 0),
                        "unit_price": float(row[2].value or 0),
                        "growth_rate": float(row[3].value or 0) / 100,
                        "introduction_year": int(row[4].value or 0),
                        "market_size": float(row[5].value or 0),
                        "price_elasticity": float(row[6].value or 0)
                    })
            
            return {
                "scenario": scenario_data,
                "equipment": equipment_data,
                "products": product_data
            }
            
        except Exception as e:
            raise ValueError(f"Failed to read template: {str(e)}")
    
    def create_scenario_from_template(self, template_file: str) -> Dict[str, Any]:
        """
        Create a new scenario from a filled template.
        
        Args:
            template_file: Path to the filled template
            
        Returns:
            Dictionary with scenario ID and analysis file path
        """
        try:
            # Read template data
            data = self.read_template(template_file)
            
            # Create scenario
            scenario = Scenario(**data["scenario"])
            self.session.add(scenario)
            self.session.flush()  # Get scenario ID
            
            # Create equipment
            for eq_data in data["equipment"]:
                equipment = Equipment(scenario_id=scenario.id, **eq_data)
                self.session.add(equipment)
            
            # Create products
            for prod_data in data["products"]:
                product = Product(scenario_id=scenario.id, **prod_data)
                self.session.add(product)
            
            # Commit changes
            self.session.commit()
            
            # Generate analysis
            output_file = f"analysis_{scenario.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            result = self.spreadsheet_service.create_scenario_dashboard(
                scenario_id=scenario.id,
                output_file=output_file,
                include_optimization=True
            )
            
            return {
                "scenario_id": scenario.id,
                "analysis_file": result.get("file_path"),
                "status": "success"
            }
            
        except Exception as e:
            self.session.rollback()
            return {"error": f"Failed to create scenario: {str(e)}"}
    
    def update_scenario_from_template(self, scenario_id: int, template_file: str) -> Dict[str, Any]:
        """
        Update an existing scenario from a filled template.
        
        Args:
            scenario_id: ID of the scenario to update
            template_file: Path to the filled template
            
        Returns:
            Dictionary with update status
        """
        try:
            # Get scenario
            scenario = self.session.query(Scenario).filter(Scenario.id == scenario_id).first()
            if not scenario:
                return {"error": f"Scenario with ID {scenario_id} not found"}
            
            # Read template data
            data = self.read_template(template_file)
            
            # Update scenario
            for key, value in data["scenario"].items():
                setattr(scenario, key, value)
            
            # Delete existing equipment and products
            self.session.query(Equipment).filter(Equipment.scenario_id == scenario_id).delete()
            self.session.query(Product).filter(Product.scenario_id == scenario_id).delete()
            
            # Create new equipment
            for eq_data in data["equipment"]:
                equipment = Equipment(scenario_id=scenario_id, **eq_data)
                self.session.add(equipment)
            
            # Create new products
            for prod_data in data["products"]:
                product = Product(scenario_id=scenario_id, **prod_data)
                self.session.add(product)
            
            # Commit changes
            self.session.commit()
            
            return {"status": "success"}
            
        except Exception as e:
            self.session.rollback()
            return {"error": f"Failed to update scenario: {str(e)}"}
    
    def create_template_from_scenario(self, scenario_id: int, output_file: str) -> Dict[str, Any]:
        """
        Create a template from an existing scenario.
        
        Args:
            scenario_id: ID of the scenario
            output_file: Path to output template file
            
        Returns:
            Dictionary with template file path
        """
        try:
            # Get scenario
            scenario = self.session.query(Scenario).filter(Scenario.id == scenario_id).first()
            if not scenario:
                return {"error": f"Scenario with ID {scenario_id} not found"}
            
            # Create template
            wb = openpyxl.Workbook()
            
            # Create Scenario Inputs sheet
            ws_scenario = wb.active
            ws_scenario.title = "Scenario Inputs"
            
            # Add scenario data
            ws_scenario.append(["Scenario Details"])
            ws_scenario.append(["Name", scenario.name])
            ws_scenario.append(["Description", scenario.description])
            ws_scenario.append(["Base Case", "Yes" if scenario.is_base_case else "No"])
            ws_scenario.append([])
            
            ws_scenario.append(["Financial Assumptions"])
            ws_scenario.append(["Initial Revenue", scenario.initial_revenue])
            ws_scenario.append(["Initial Costs", scenario.initial_costs])
            ws_scenario.append(["Annual Revenue Growth (%)", scenario.annual_revenue_growth * 100])
            ws_scenario.append(["Annual Cost Growth (%)", scenario.annual_cost_growth * 100])
            ws_scenario.append(["Debt Ratio (%)", scenario.debt_ratio * 100])
            ws_scenario.append(["Interest Rate (%)", scenario.interest_rate * 100])
            ws_scenario.append(["Tax Rate (%)", scenario.tax_rate * 100])
            
            # Create Equipment Inputs sheet
            ws_equipment = wb.create_sheet("Equipment Inputs")
            headers = ["Name", "Cost", "Useful Life", "Max Capacity", 
                      "Maintenance Cost (%)", "Availability (%)", 
                      "Purchase Year", "Leased", "Lease Rate (%)", "Lease Type"]
            ws_equipment.append(headers)
            
            # Add equipment data
            equipment = self.session.query(Equipment).filter(Equipment.scenario_id == scenario_id).all()
            for eq in equipment:
                ws_equipment.append([
                    eq.name, eq.cost, eq.useful_life, eq.max_capacity,
                    eq.maintenance_cost_pct * 100, eq.availability_pct * 100,
                    eq.purchase_year, "Yes" if eq.is_leased else "No",
                    eq.lease_rate * 100 if eq.lease_rate else "",
                    eq.lease_type
                ])
            
            # Create Product Inputs sheet
            ws_product = wb.create_sheet("Product Inputs")
            headers = ["Name", "Initial Units", "Unit Price", "Growth Rate (%)",
                      "Introduction Year", "Market Size", "Price Elasticity"]
            ws_product.append(headers)
            
            # Add product data
            products = self.session.query(Product).filter(Product.scenario_id == scenario_id).all()
            for prod in products:
                ws_product.append([
                    prod.name, prod.initial_units, prod.unit_price,
                    prod.growth_rate * 100, prod.introduction_year,
                    prod.market_size, prod.price_elasticity
                ])
            
            # Apply formatting
            for ws in [ws_scenario, ws_equipment, ws_product]:
                for row in ws.iter_rows():
                    for cell in row:
                        cell.font = Font(name="Calibri")
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
            
            # Format headers
            for ws in [ws_scenario, ws_equipment, ws_product]:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Save template
            wb.save(output_file)
            
            return {"status": "success", "file_path": output_file}
            
        except Exception as e:
            return {"error": f"Failed to create template: {str(e)}"} 