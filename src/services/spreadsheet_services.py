"""
Spreadsheet integration service for the manufacturing expansion financial model.

This module provides capabilities for exporting data to and importing data from
Excel and Google Sheets, enabling these applications to serve as a user interface.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
import os
from pathlib import Path
from sqlalchemy.exc import SQLAlchemyError
from openpyxl.exceptions import InvalidFileException
from google.auth.exceptions import GoogleAuthError
from googleapiclient.errors import HttpError as GoogleAPIError

# Excel dependencies
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Google Sheets dependencies
try:
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    GOOGLE_SHEETS_AVAILABLE = True
except ImportError:
    GOOGLE_SHEETS_AVAILABLE = False

from advanced_pmodeler.models import (
    Scenario, Equipment, Product, CostDriver, FinancialProjection,
    get_session
)

class SpreadsheetService:
    """
    Service class for spreadsheet integration.
    
    This class provides methods for:
    - Exporting data to Excel and Google Sheets
    - Importing data from Excel and Google Sheets
    - Creating formatted reports and dashboards
    - Managing spreadsheet templates
    """
    
    def __init__(self, session):
        """
        Initialize SpreadsheetService with a database session.
        
        Args:
            session: SQLAlchemy database session
        """
        self.session = session
        self._google_sheets_service = None
    
    def export_to_excel(self, data: Union[pd.DataFrame, Dict[str, Any]], 
                       output_file: str,
                       sheet_name: str = "Sheet1",
                       template_file: Optional[str] = None) -> Dict[str, Any]:
        """
        Export data to Excel with optional template.
        
        Args:
            data: DataFrame or dictionary to export
            output_file: Path to output Excel file
            sheet_name: Name of the sheet to create
            template_file: Optional template file to use
            
        Returns:
            Dictionary with export status and file path
        """
        if not EXCEL_AVAILABLE:
            return {"error": "Excel support not available. Please install openpyxl."}
        
        try:
            # Convert dictionary to DataFrame if needed
            if isinstance(data, dict):
                data = pd.DataFrame(data)
            
            # Create workbook
            if template_file and os.path.exists(template_file):
                wb = openpyxl.load_workbook(template_file)
            else:
                wb = openpyxl.Workbook()
            
            # Get or create sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Write data
            for row in data.itertuples():
                ws.append(row)
            
            # Apply formatting
            self._apply_excel_formatting(ws, data)
            
            # Save file
            wb.save(output_file)
            
            return {"status": "success", "file_path": output_file}
            
        except InvalidFileException as e:
            return {"error": f"Invalid Excel file: {str(e)}"}
        except PermissionError as e:
            return {"error": f"Permission denied when saving Excel file: {str(e)}"}
        except ValueError as e:
            return {"error": f"Invalid data format: {str(e)}"}
        except Exception as e:
            return {"error": f"Failed to export to Excel: {str(e)}"}
    
    def export_to_google_sheets(self, data: Union[pd.DataFrame, Dict[str, Any]],
                              spreadsheet_id: str,
                              range_name: str,
                              credentials_path: str) -> Dict[str, Any]:
        """
        Export data to Google Sheets.
        
        Args:
            data: DataFrame or dictionary to export
            spreadsheet_id: ID of the Google Sheet
            range_name: Range to write to (e.g., "Sheet1!A1")
            credentials_path: Path to Google credentials file
            
        Returns:
            Dictionary with export status
        """
        if not GOOGLE_SHEETS_AVAILABLE:
            return {"error": "Google Sheets support not available. Please install required packages."}
        
        try:
            # Get Google Sheets service
            service = self._get_google_sheets_service(credentials_path)
            
            # Convert dictionary to DataFrame if needed
            if isinstance(data, dict):
                data = pd.DataFrame(data)
            
            # Convert DataFrame to values
            values = data.values.tolist()
            
            # Write data
            body = {
                'values': values
            }
            
            service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=range_name,
                valueInputOption='RAW',
                body=body
            ).execute()
            
            return {"status": "success"}
            
        except GoogleAuthError as e:
            return {"error": f"Google authentication error: {str(e)}"}
        except GoogleAPIError as e:
            return {"error": f"Google Sheets API error: {str(e)}"}
        except ValueError as e:
            return {"error": f"Invalid data format: {str(e)}"}
        except Exception as e:
            return {"error": f"Failed to export to Google Sheets: {str(e)}"}
    
    def create_scenario_dashboard(self, scenario_id: int,
                                output_file: str,
                                include_optimization: bool = False) -> Dict[str, Any]:
        """
        Create a comprehensive Excel dashboard for a scenario.
        
        Args:
            scenario_id: ID of the scenario
            output_file: Path to output Excel file
            include_optimization: Whether to include optimization results
            
        Returns:
            Dictionary with export status and file path
        """
        if not EXCEL_AVAILABLE:
            return {"error": "Excel support not available. Please install openpyxl."}
        
        try:
            # Get scenario data
            scenario = self.session.query(Scenario).filter(Scenario.id == scenario_id).first()
            if not scenario:
                return {"error": f"Scenario with ID {scenario_id} not found"}
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Create sheets
            sheets = {
                "Overview": self._create_overview_sheet(wb, scenario),
                "Financial Projections": self._create_financial_sheet(wb, scenario),
                "Equipment Analysis": self._create_equipment_sheet(wb, scenario),
                "Product Analysis": self._create_product_sheet(wb, scenario)
            }
            
            if include_optimization:
                sheets["Optimization"] = self._create_optimization_sheet(wb, scenario)
            
            # Save file
            wb.save(output_file)
            
            return {"status": "success", "file_path": output_file}
            
        except SQLAlchemyError as e:
            return {"error": f"Database error while creating dashboard: {str(e)}"}
        except InvalidFileException as e:
            return {"error": f"Invalid Excel file: {str(e)}"}
        except PermissionError as e:
            return {"error": f"Permission denied when saving Excel file: {str(e)}"}
        except Exception as e:
            return {"error": f"Failed to create dashboard: {str(e)}"}
    
    def _apply_excel_formatting(self, ws, data: pd.DataFrame):
        """Apply formatting to Excel worksheet."""
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Format data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
    
    def _get_google_sheets_service(self, credentials_path: str):
        """Get or create Google Sheets service."""
        if self._google_sheets_service is None:
            creds = None
            if os.path.exists('token.json'):
                creds = Credentials.from_authorized_user_file('token.json')
            
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                else:
                    flow = InstalledAppFlow.from_client_secrets_file(
                        credentials_path, 
                        ['https://www.googleapis.com/auth/spreadsheets']
                    )
                    creds = flow.run_local_server(port=0)
                
                with open('token.json', 'w') as token:
                    token.write(creds.to_json())
            
            self._google_sheets_service = build('sheets', 'v4', credentials=creds)
        
        return self._google_sheets_service
    
    def _create_overview_sheet(self, wb, scenario: Scenario) -> openpyxl.worksheet.worksheet.Worksheet:
        """Create overview sheet with scenario details."""
        ws = wb.active
        ws.title = "Overview"
        
        # Add scenario details
        ws.append(["Scenario Overview"])
        ws.append(["Name", scenario.name])
        ws.append(["Description", scenario.description])
        ws.append(["Base Case", "Yes" if scenario.is_base_case else "No"])
        ws.append([])
        
        # Add financial assumptions
        ws.append(["Financial Assumptions"])
        ws.append(["Initial Revenue", scenario.initial_revenue])
        ws.append(["Initial Costs", scenario.initial_costs])
        ws.append(["Annual Revenue Growth", f"{scenario.annual_revenue_growth*100}%"])
        ws.append(["Annual Cost Growth", f"{scenario.annual_cost_growth*100}%"])
        ws.append(["Debt Ratio", f"{scenario.debt_ratio*100}%"])
        ws.append(["Interest Rate", f"{scenario.interest_rate*100}%"])
        ws.append(["Tax Rate", f"{scenario.tax_rate*100}%"])
        
        return ws
    
    def _create_financial_sheet(self, wb, scenario: Scenario) -> openpyxl.worksheet.worksheet.Worksheet:
        """Create financial projections sheet."""
        ws = wb.create_sheet("Financial Projections")
        
        # Get financial projections
        projections = self.session.query(FinancialProjection).filter(
            FinancialProjection.scenario_id == scenario.id
        ).order_by(FinancialProjection.year).all()
        
        if not projections:
            ws.append(["No financial projections available"])
            return ws
        
        # Add headers
        headers = ["Year", "Revenue", "COGS", "Gross Profit", "Operating Expenses", 
                  "EBITDA", "Depreciation", "EBIT", "Interest", "Tax", "Net Income"]
        ws.append(headers)
        
        # Add data
        for p in projections:
            ws.append([
                p.year, p.revenue, p.cogs, p.gross_profit, p.operating_expenses,
                p.ebitda, p.depreciation, p.ebit, p.interest, p.tax, p.net_income
            ])
        
        return ws
    
    def _create_equipment_sheet(self, wb, scenario: Scenario) -> openpyxl.worksheet.worksheet.Worksheet:
        """Create equipment analysis sheet."""
        ws = wb.create_sheet("Equipment Analysis")
        
        # Get equipment
        equipment = self.session.query(Equipment).filter(
            Equipment.scenario_id == scenario.id
        ).all()
        
        if not equipment:
            ws.append(["No equipment data available"])
            return ws
        
        # Add headers
        headers = ["Name", "Cost", "Useful Life", "Max Capacity", "Maintenance Cost %",
                  "Availability %", "Purchase Year", "Leased", "Lease Rate", "Lease Type"]
        ws.append(headers)
        
        # Add data
        for e in equipment:
            ws.append([
                e.name, e.cost, e.useful_life, e.max_capacity,
                f"{e.maintenance_cost_pct*100}%", f"{e.availability_pct*100}%",
                e.purchase_year, "Yes" if e.is_leased else "No",
                f"{e.lease_rate*100}%" if e.lease_rate else "",
                e.lease_type if e.lease_type else ""
            ])
        
        return ws
    
    def _create_product_sheet(self, wb, scenario: Scenario) -> openpyxl.worksheet.worksheet.Worksheet:
        """Create product analysis sheet."""
        ws = wb.create_sheet("Product Analysis")
        
        # Get products
        products = self.session.query(Product).filter(
            Product.scenario_id == scenario.id
        ).all()
        
        if not products:
            ws.append(["No product data available"])
            return ws
        
        # Add headers
        headers = ["Name", "Initial Units", "Unit Price", "Growth Rate",
                  "Introduction Year", "Market Size", "Price Elasticity"]
        ws.append(headers)
        
        # Add data
        for p in products:
            ws.append([
                p.name, p.initial_units, p.unit_price,
                f"{p.growth_rate*100}%", p.introduction_year,
                p.market_size, p.price_elasticity
            ])
        
        return ws
    
    def _create_optimization_sheet(self, wb, scenario: Scenario) -> openpyxl.worksheet.worksheet.Worksheet:
        """Create optimization results sheet."""
        ws = wb.create_sheet("Optimization")
        
        # Get optimization metadata
        if not hasattr(scenario, 'optimization_metadata') or not scenario.optimization_metadata:
            ws.append(["No optimization data available"])
            return ws
        
        metadata = scenario.optimization_metadata
        
        # Add optimization details
        ws.append(["Optimization Results"])
        ws.append(["Last Optimized", metadata.get("last_optimized", "N/A")])
        ws.append([])
        
        # Add optimization results
        if "optimization_results" in metadata:
            results = metadata["optimization_results"]
            ws.append(["Equipment Purchase Plan"])
            
            # Add equipment purchase plan
            if "equipment_purchase_plan" in results:
                ws.append(["Year", "Equipment", "Cost", "Capacity Added"])
                for year, purchases in results["equipment_purchase_plan"].items():
                    for purchase in purchases:
                        ws.append([
                            year,
                            purchase["equipment_name"],
                            purchase["cost"],
                            purchase["capacity_added"]
                        ])
            
            # Add expected improvements
            if "expected_improvements" in results:
                ws.append([])
                ws.append(["Expected Utilization Improvements"])
                ws.append(["Equipment", "Current Utilization", "Expected Utilization", "Improvement"])
                for eq_id, improvement in results["expected_improvements"].items():
                    ws.append([
                        eq_id,
                        f"{improvement['current_utilization']}%",
                        f"{improvement['expected_utilization']}%",
                        f"{improvement['improvement']}%"
                    ])
        
        return ws 